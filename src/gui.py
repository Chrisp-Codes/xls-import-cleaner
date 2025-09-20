import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell
import tkinter as tk
from tkinter import filedialog, messagebox


def clean_import_file(file_path: str, mode: str = "startdaten"):
    wb = load_workbook(file_path)
    ws = wb.active

    # Header finden
    header_row = None
    for row in range(1, ws.max_row + 1):
        if str(ws[f"A{row}"].value).strip() == "Personalnummer":
            header_row = row
            break

    if header_row is None:
        raise ValueError("Kein Header mit 'Personalnummer' in Spalte A gefunden.")

    header_keep_start = max(1, header_row - 2)

    # Zeilen 5 bis header_keep_start-1 löschen
    for row in range(5, header_keep_start):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            if isinstance(cell, MergedCell):
                continue
            cell.value = None

    # Spaltenauswahl
    keep_cols = ["C", "E"]  # Vorname (C), Nachname (E)
    if mode == "startdaten":
        keep_cols.append("BL")
    elif mode == "startwerte":
        keep_cols.extend([get_column_letter(c) for c in range(65, 72)])  # BM-BS

    # Datenbereich ab header_row+1
    for row in range(header_row + 1, ws.max_row + 1):
        if all(ws.cell(row=row, column=col).value is None for col in range(1, ws.max_column + 1)):
            break
        for col in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col)
            cell = ws.cell(row=row, column=col)
            if isinstance(cell, MergedCell):
                continue
            if col_letter not in keep_cols:
                cell.value = None

    # Neue Datei speichern
    base, ext = os.path.splitext(file_path)
    new_file = f"{base}_bereinigt{ext}"
    wb.save(new_file)
    return new_file


# --- Tkinter GUI ---
def run_gui():
    def process_file():
        mode = mode_var.get()
        if mode not in ["startwerte", "startdaten"]:
            messagebox.showerror("Fehler", "Bitte wähle eine Option (Startwerte oder Startdaten).")
            return

        file_path = filedialog.askopenfilename(
            title="Importdatei auswählen",
            filetypes=[("Excel Dateien", "*.xlsx *.xls")]
        )
        if not file_path:
            return

        try:
            new_file = clean_import_file(file_path, mode=mode)
            messagebox.showinfo("Fertig", f"Datei bereinigt gespeichert:\n{new_file}")
            root.destroy()  # Fenster schließen nach Erfolg
        except Exception as e:
            messagebox.showerror("Fehler", str(e))

    root = tk.Tk()
    root.title("XLS Import Cleaner")
    root.geometry("450x200")

    tk.Label(root, text="Was möchtest du reimportieren?", font=("Arial", 12)).pack(pady=10)

    mode_var = tk.StringVar()
    mode_var.set(None)

    # Frame für Radiobuttons nebeneinander
    frame = tk.Frame(root)
    frame.pack(pady=5)

    tk.Radiobutton(frame, text="Startwerte", variable=mode_var, value="startwerte", font=("Arial", 11)).pack(side="left", padx=20)
    tk.Radiobutton(frame, text="Startdaten", variable=mode_var, value="startdaten", font=("Arial", 11)).pack(side="left", padx=20)

    tk.Button(root, text="Datei auswählen und bereinigen", command=process_file, font=("Arial", 11)).pack(pady=20)

    root.mainloop()


if __name__ == "__main__":
    run_gui()
